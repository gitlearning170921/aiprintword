# -*- coding: utf-8 -*-
"""
签字位真实版式分析。

输入：文档路径 + 已识别角色列表。
输出：基于真实表格/段落结构的版式判定，含每个角色的：
  - 姓名签字单元格位置 (table/row/col 或 paragraph index)
  - 日期签字单元格位置
  - 角色与日期的相对位置（同格/右方/下方/段落内联）
  - 分隔方式（`/`、空格、空单元格、单元格分隔、换行）
  - 多角色排列轴（左到右 horizontal / 上到下 vertical / 混合 mixed）
"""
from __future__ import annotations

import os
import re
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

from sign_handlers.config import ROLE_ID_TO_KEYWORD, role_keywords
from sign_handlers.detect_fields import (
    _cell_role_ids_multiline,
    _iter_docx_tables_for_detect,
    _norm,
)
from sign_handlers.docx_revision_text import (
    cell_effective_text,
    paragraph_effective_text,
)


_DATE_VALUE_RE = re.compile(
    r"(?:\d{4}\s*[年./\-]\s*\d{1,2}\s*[月./\-]\s*\d{1,2}\s*日?"
    r"|\d{4}[./\-]\d{1,2}[./\-]\d{1,2}"
    r"|\d{1,2}[./\-]\d{1,2}[./\-]\d{2,4}"
    r"|[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4})"
)
_DATE_LABEL_RE = re.compile(
    r"(日\s*期|签署日期|签字日期|填报日期|年\s*月\s*日|Date|Signed\s*Date|Sign\s*Date)",
    re.IGNORECASE,
)


def _norm_keep_newlines(s: str) -> str:
    """版式分析保留换行，便于判定「日期在角色下方/换行分隔」。"""
    if not s:
        return ""
    lines = re.split(r"[\r\n]+", str(s))
    lines = [re.sub(r"[ \t\u00A0\u3000]+", " ", ln).strip() for ln in lines]
    lines = [ln for ln in lines if ln]
    return "\n".join(lines)


def _has_date_token(text: str) -> bool:
    s = text or ""
    if not s:
        return False
    return bool(_DATE_LABEL_RE.search(s) or _DATE_VALUE_RE.search(s))


def _looks_like_date_cell(text: str) -> bool:
    """单元格是否“像专门放日期的格”。"""
    t = (text or "").strip()
    if not t:
        return False
    if len(t) > 64:
        return False
    return _has_date_token(t)


def _looks_like_empty_date_slot(text: str) -> bool:
    """空格或仅含「年月日 / ___ / xxxx-xx-xx」占位符的格。"""
    t = (text or "").strip()
    if not t:
        return True
    if len(t) > 24:
        return False
    if re.fullmatch(r"[年月日\s_\-/.:：]+", t):
        return True
    return False


def _role_keywords_for(rid: str) -> List[str]:
    try:
        return list(role_keywords(rid) or [])
    except Exception:
        return list(ROLE_ID_TO_KEYWORD.get(rid, []) or [])


def _analyze_same_cell(cell_text: str, role_kws: Sequence[str]) -> Tuple[str, str]:
    """单元格内同时含角色关键词 + 日期时，返回 (separator, position)。

    position 取值：
      - "right"  ：日期在角色文本之后（同一行视觉上在右侧）
      - "left"   ：日期在角色文本之前
      - "below"  ：日期与角色之间存在换行
      - "inline" ：兜底
    """
    s = cell_text or ""
    if not s:
        return ("none", "none")
    role_pos = -1
    role_end = 0
    for kw in sorted([k for k in role_kws if k], key=len, reverse=True):
        i = s.find(kw)
        if i >= 0:
            role_pos = i
            role_end = i + len(kw)
            break
    if role_pos < 0:
        return ("unknown", "inline")
    m = _DATE_VALUE_RE.search(s) or _DATE_LABEL_RE.search(s)
    if not m:
        return ("none", "none")
    if m.start() < role_pos:
        between = s[m.end():role_pos]
        position = "left"
    else:
        between = s[role_end:m.start()]
        position = "right"
    between_stripped = between.strip()
    has_newline = "\n" in between or "\r" in between
    if has_newline:
        position = "below"
    if between == "":
        sep = "adjacent"
    elif has_newline and not between_stripped:
        sep = "newline"
    elif not between_stripped:
        sep = "space"
    elif "/" in between_stripped or "／" in between_stripped:
        sep = "slash"
    elif "\\" in between_stripped:
        sep = "backslash"
    elif has_newline:
        sep = "newline"
    elif re.fullmatch(r"[\s:：,\-、]+", between_stripped):
        sep = "punct:" + between_stripped[:6]
    else:
        sep = "other:" + between_stripped[:6]
    return (sep, position)


def _build_table_grid(table) -> List[List[Dict[str, Any]]]:
    grid: List[List[Dict[str, Any]]] = []
    for row in table.rows:
        row_cells = []
        for cell in row.cells:
            txt = _norm_keep_newlines(cell_effective_text(cell) or "")
            row_cells.append({"text": txt})
        grid.append(row_cells)
    return grid


def _find_role_primary_cell(
    grid: List[List[Dict[str, Any]]], role_id: str
) -> Optional[Tuple[int, int, str]]:
    """选最像「标签」的格：包含该角色关键词、且文本最短（避免命中长正文段）。"""
    best: Optional[Tuple[int, int, int, str]] = None
    for ri, row in enumerate(grid):
        for ci, cell in enumerate(row):
            txt = cell["text"]
            if not txt or len(txt) > 96:
                continue
            ids = _cell_role_ids_multiline(txt)
            if role_id not in ids:
                continue
            key = (len(txt), ri, ci)
            if best is None or key < (best[0], best[1], best[2]):
                best = (len(txt), ri, ci, txt)
    if best is None:
        return None
    return (best[1], best[2], best[3])


def _detect_role_date_relation(
    grid: List[List[Dict[str, Any]]],
    ri: int,
    ci: int,
    role_text: str,
    role_kws: Sequence[str],
) -> Dict[str, Any]:
    H = len(grid)
    W = max((len(r) for r in grid), default=0)

    if _has_date_token(role_text):
        sep, pos = _analyze_same_cell(role_text, role_kws)
        return {
            "relation": "same_cell",
            "position": pos,
            "separator": sep,
            "date_row": ri,
            "date_col": ci,
        }

    if ci + 1 < W:
        right_txt = grid[ri][ci + 1]["text"]
        if _looks_like_date_cell(right_txt) or _looks_like_empty_date_slot(right_txt):
            return {
                "relation": "different_cell",
                "position": "right",
                "separator": "empty_cell" if not right_txt.strip() else "cell",
                "date_row": ri,
                "date_col": ci + 1,
            }

    if ri + 1 < H and ci < len(grid[ri + 1]):
        below_txt = grid[ri + 1][ci]["text"]
        if _looks_like_date_cell(below_txt) or _looks_like_empty_date_slot(below_txt):
            return {
                "relation": "different_cell",
                "position": "below",
                "separator": "empty_cell" if not below_txt.strip() else "cell",
                "date_row": ri + 1,
                "date_col": ci,
            }

    for dc in range(2, min(5, W - ci)):
        right_txt = grid[ri][ci + dc]["text"]
        if _looks_like_date_cell(right_txt):
            return {
                "relation": "different_cell",
                "position": "right",
                "separator": "cell",
                "date_row": ri,
                "date_col": ci + dc,
            }

    for dr in range(2, min(4, H - ri)):
        if ci >= len(grid[ri + dr]):
            continue
        below_txt = grid[ri + dr][ci]["text"]
        if _looks_like_date_cell(below_txt):
            return {
                "relation": "different_cell",
                "position": "below",
                "separator": "cell",
                "date_row": ri + dr,
                "date_col": ci,
            }

    return {
        "relation": "none",
        "position": "none",
        "separator": "none",
        "date_row": None,
        "date_col": None,
    }


def _determine_axis(role_positions: List[Tuple[int, int]]) -> str:
    if not role_positions:
        return "unknown"
    if len(role_positions) == 1:
        return "single"
    rows = {p[0] for p in role_positions}
    cols = {p[1] for p in role_positions}
    if len(rows) == 1 and len(cols) >= 2:
        return "horizontal"
    if len(cols) == 1 and len(rows) >= 2:
        return "vertical"
    if len(rows) >= 2 and len(cols) >= 2:
        # 倾向：若同一行存在多个角色 -> 水平为主；否则垂直为主
        row_count = {}
        col_count = {}
        for r, c in role_positions:
            row_count[r] = row_count.get(r, 0) + 1
            col_count[c] = col_count.get(c, 0) + 1
        max_per_row = max(row_count.values())
        max_per_col = max(col_count.values())
        if max_per_row > max_per_col:
            return "horizontal"
        if max_per_col > max_per_row:
            return "vertical"
        return "mixed"
    return "unknown"


def _analyze_paragraph_layout(doc, role_set: Set[str]) -> Dict[str, Any]:
    role_layouts: Dict[str, Any] = {}
    for pi, p in enumerate(doc.paragraphs):
        t = _norm_keep_newlines(paragraph_effective_text(p) or "")
        if not t:
            continue
        ids = set(_cell_role_ids_multiline(t)) & role_set
        if not ids:
            continue
        for rid in ids:
            if rid in role_layouts:
                continue
            has_date = _has_date_token(t)
            if has_date:
                sep, pos = _analyze_same_cell(t, _role_keywords_for(rid))
            else:
                sep, pos = ("none", "none")
            role_layouts[rid] = {
                "name_slot": True,
                "name_loc": f"paragraph#{pi + 1}",
                "date_slot": has_date,
                "date_relation": "paragraph_inline" if has_date else "none",
                "date_position": pos if has_date else "none",
                "separator": sep,
                "date_loc": f"paragraph#{pi + 1}" if has_date else None,
            }
    return {
        "ok": bool(role_layouts),
        "kind": "docx_paragraph",
        "arrangement": "inline" if role_layouts else "unknown",
        "role_layouts": role_layouts,
    }


def analyze_docx_layout(path: str, role_ids: Sequence[str]) -> Dict[str, Any]:
    from docx import Document

    role_set: Set[str] = {str(x).strip() for x in (role_ids or []) if str(x).strip()}
    role_set = {rid for rid in role_set if rid in ROLE_ID_TO_KEYWORD}
    if not role_set:
        return {"ok": False, "error": "no roles", "role_layouts": {}}
    doc = Document(path)

    best_table: Optional[Dict[str, Any]] = None
    for ti, table in enumerate(
        _iter_docx_tables_for_detect(doc, max_body_tables=32, max_total=64)
    ):
        grid = _build_table_grid(table)
        if not grid:
            continue
        primaries: Dict[str, Tuple[int, int, str]] = {}
        for rid in role_set:
            cell = _find_role_primary_cell(grid, rid)
            if cell:
                primaries[rid] = cell
        if not primaries:
            continue
        if best_table is None or len(primaries) > best_table["hit_count"]:
            best_table = {
                "ti": ti,
                "grid": grid,
                "primaries": primaries,
                "hit_count": len(primaries),
            }
            if len(primaries) == len(role_set):
                break

    if not best_table:
        return _analyze_paragraph_layout(doc, role_set)

    grid = best_table["grid"]
    primaries = best_table["primaries"]
    ti = best_table["ti"]

    role_layouts: Dict[str, Any] = {}
    positions: List[Tuple[int, int]] = []
    for rid, (ri, ci, txt) in primaries.items():
        rel = _detect_role_date_relation(grid, ri, ci, txt, _role_keywords_for(rid))
        role_layouts[rid] = {
            "name_slot": True,
            "name_loc": f"table#{ti + 1}.r{ri + 1}.c{ci + 1}",
            "date_slot": rel["relation"] != "none",
            "date_relation": rel["relation"],
            "date_position": rel["position"],
            "separator": rel["separator"],
            "date_loc": (
                f"table#{ti + 1}.r{rel['date_row'] + 1}.c{rel['date_col'] + 1}"
                if rel["date_row"] is not None
                else None
            ),
        }
        positions.append((ri, ci))

    axis = _determine_axis(positions)

    # 表格内未命中的角色，补段落 fallback
    missing = role_set - set(role_layouts.keys())
    if missing:
        para_layout = _analyze_paragraph_layout(doc, missing)
        for rid, info in (para_layout.get("role_layouts") or {}).items():
            role_layouts[rid] = info

    return {
        "ok": True,
        "kind": "docx_table",
        "table_index": ti,
        "arrangement": axis,
        "role_layouts": role_layouts,
    }


def analyze_xlsx_layout(path: str, role_ids: Sequence[str]) -> Dict[str, Any]:
    from openpyxl import load_workbook

    role_set: Set[str] = {str(x).strip() for x in (role_ids or []) if str(x).strip()}
    role_set = {rid for rid in role_set if rid in ROLE_ID_TO_KEYWORD}
    if not role_set:
        return {"ok": False, "error": "no roles", "role_layouts": {}}

    wb = load_workbook(path, data_only=True)

    best: Optional[Dict[str, Any]] = None
    for ws in wb.worksheets:
        max_r = min(int(ws.max_row or 0), 300)
        max_c = min(int(ws.max_column or 0), 60)
        if max_r <= 0 or max_c <= 0:
            continue
        grid: List[List[Dict[str, Any]]] = []
        for r in range(1, max_r + 1):
            row_cells = []
            for c in range(1, max_c + 1):
                v = ws.cell(row=r, column=c).value
                row_cells.append(
                    {
                        "text": _norm_keep_newlines(str(v)) if v is not None else ""
                    }
                )
            grid.append(row_cells)
        primaries: Dict[str, Tuple[int, int, str]] = {}
        for rid in role_set:
            cell = _find_role_primary_cell(grid, rid)
            if cell:
                primaries[rid] = cell
        if not primaries:
            continue
        if best is None or len(primaries) > best["hit_count"]:
            best = {
                "sheet": ws.title or "",
                "grid": grid,
                "primaries": primaries,
                "hit_count": len(primaries),
            }
            if len(primaries) == len(role_set):
                break

    if not best:
        return {
            "ok": False,
            "kind": "xlsx",
            "role_layouts": {},
            "error": "no role cells found",
        }

    grid = best["grid"]
    primaries = best["primaries"]
    sheet = best["sheet"]

    role_layouts: Dict[str, Any] = {}
    positions: List[Tuple[int, int]] = []
    for rid, (ri, ci, txt) in primaries.items():
        rel = _detect_role_date_relation(grid, ri, ci, txt, _role_keywords_for(rid))
        role_layouts[rid] = {
            "name_slot": True,
            "name_loc": f"{sheet}!r{ri + 1}.c{ci + 1}",
            "date_slot": rel["relation"] != "none",
            "date_relation": rel["relation"],
            "date_position": rel["position"],
            "separator": rel["separator"],
            "date_loc": (
                f"{sheet}!r{rel['date_row'] + 1}.c{rel['date_col'] + 1}"
                if rel["date_row"] is not None
                else None
            ),
        }
        positions.append((ri, ci))

    return {
        "ok": True,
        "kind": "xlsx",
        "sheet": sheet,
        "arrangement": _determine_axis(positions),
        "role_layouts": role_layouts,
    }


def _layout_from_document_rule(
    rule: Dict[str, Any], role_ids: Sequence[str]
) -> Dict[str, Any]:
    """按 document_layout_rules 人工登记生成 signature_layout。"""
    from sign_handlers.config import ROLE_ID_TO_KEYWORD

    role_set = {str(x).strip() for x in (role_ids or []) if str(x).strip()}
    role_set = {rid for rid in role_set if rid in ROLE_ID_TO_KEYWORD}
    rel = str(rule.get("date_relation") or "none").strip() or "none"
    pos = str(rule.get("date_position") or "none").strip() or "none"
    sep = str(rule.get("separator") or "none").strip() or "none"
    has_date = rel not in ("", "none")
    role_layouts: Dict[str, Any] = {}
    for rid in role_set:
        role_layouts[rid] = {
            "name_slot": True,
            "date_slot": has_date,
            "date_relation": rel,
            "date_position": pos,
            "separator": sep,
            "name_loc": "document_layout_rule",
            "date_loc": "document_layout_rule" if has_date else None,
        }
    return {
        "ok": True,
        "kind": "document_layout_rule",
        "arrangement": str(rule.get("arrangement") or "unknown"),
        "role_layouts": role_layouts,
        "source": "document_layout_rule",
        "pattern": str(rule.get("pattern") or ""),
    }


def analyze_signature_layout(
    path: str, ext: str, role_ids: Sequence[str], source_name: str = ""
) -> Dict[str, Any]:
    e = (ext or "").lower()
    try:
        detected: Dict[str, Any] = {}
        if e == ".docx":
            detected = analyze_docx_layout(path, role_ids)
        elif e == ".xlsx":
            detected = analyze_xlsx_layout(path, role_ids)
        else:
            return {
                "ok": False,
                "error": f"unsupported ext: {ext}",
                "role_layouts": {},
            }
        # 优先真实文档结构分析；只有失败时才回退文件名规则。
        if isinstance(detected, dict) and detected.get("ok"):
            return detected
        if source_name:
            try:
                from sign_handlers.detect_correction_slot_rules import (
                    match_document_layout_rule,
                )

                doc_rule = match_document_layout_rule(source_name)
                if doc_rule and isinstance(doc_rule, dict):
                    laid = _layout_from_document_rule(doc_rule, role_ids)
                    if laid.get("ok"):
                        laid["fallback_reason"] = str(
                            (detected or {}).get("error") or "layout_detect_not_ok"
                        )[:240]
                        return laid
            except Exception:
                pass
        return detected
    except Exception as exc:
        return {"ok": False, "error": str(exc), "role_layouts": {}}
