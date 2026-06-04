# -*- coding: utf-8 -*-
"""导出与自动签字落位一致缩放比例的 PNG，供工作台手动贴图。"""
from __future__ import annotations

import io
import zipfile
from typing import Any, Dict, List, Optional, Tuple

try:
    from PIL import Image
except Exception:  # pragma: no cover
    Image = None  # type: ignore

from sign_handlers.png_word_compat import (
    prepare_png_for_word,
    prepare_signature_date_pair_for_word,
)

# 与 sign_docx._PIC_WIDTH 一致（Cm(2.8)）
DOCX_INSERT_WIDTH_CM = 2.8
# Word 粘贴/显示常用 DPI；与 Cm(2.8) 对应的像素宽度约 106px
DOCX_INSERT_DPI = 96

MATERIAL_SHORT_TAGS: Dict[Tuple[str, str], str] = {
    ("author", "sig"): "编签",
    ("author", "date"): "编日",
    ("reviewer", "sig"): "审签",
    ("reviewer", "date"): "审日",
    ("approver", "sig"): "批签",
    ("approver", "date"): "批日",
}

_EXPORT_ROLE_ORDER = ("author", "reviewer", "approver")
_EXPORT_KIND_ORDER = ("sig", "date")


def scale_png_for_docx_insert(
    png_bytes: Optional[bytes],
    *,
    width_cm: float = DOCX_INSERT_WIDTH_CM,
    dpi: int = DOCX_INSERT_DPI,
) -> Optional[bytes]:
    """将 PNG 缩放到与 Word add_picture(width=Cm(2.8)) 一致的显示像素尺寸。"""
    if not png_bytes or Image is None:
        return png_bytes
    try:
        im = Image.open(io.BytesIO(png_bytes))
    except Exception:
        return png_bytes
    w, h = im.size
    if w < 1 or h < 1:
        return png_bytes
    target_w = max(1, int(round(float(width_cm) / 2.54 * float(dpi))))
    target_h = max(1, int(round(h * (target_w / float(w)))))
    if im.mode not in ("RGB", "RGBA"):
        im = im.convert("RGBA")
    flat = Image.new("RGBA", (w, h), (255, 255, 255, 255))
    flat.alpha_composite(im)
    rgb = flat.convert("RGB").resize((target_w, target_h), Image.Resampling.LANCZOS)
    out = io.BytesIO()
    rgb.save(out, format="PNG", dpi=(dpi, dpi), optimize=True)
    return out.getvalue()


def scale_png_to_pixel_box(
    png_bytes: Optional[bytes],
    target_w: int,
    target_h: int,
) -> Optional[bytes]:
    """等比缩放 PNG 到目标像素框（与 Excel 落位 _fit_to_box_excel 结果一致）。"""
    if not png_bytes or Image is None:
        return png_bytes
    tw = max(1, int(target_w))
    th = max(1, int(target_h))
    try:
        im = Image.open(io.BytesIO(png_bytes))
    except Exception:
        return png_bytes
    w, h = im.size
    if w < 1 or h < 1:
        return png_bytes
    if im.mode not in ("RGB", "RGBA"):
        im = im.convert("RGBA")
    flat = Image.new("RGBA", (w, h), (255, 255, 255, 255))
    flat.alpha_composite(im)
    rgb = flat.convert("RGB").resize((tw, th), Image.Resampling.LANCZOS)
    out = io.BytesIO()
    rgb.save(out, format="PNG", optimize=True)
    return out.getvalue()


def _coerce_png_bytes(val) -> Optional[bytes]:
    if val is None:
        return None
    if isinstance(val, memoryview):
        val = val.tobytes()
    elif isinstance(val, bytearray):
        val = bytes(val)
    elif not isinstance(val, bytes):
        try:
            val = bytes(val)
        except Exception:
            return None
    return val if len(val) > 0 else None


def _role_map_entry_nonempty(pair: dict, *, mysql_store=None) -> bool:
    if not isinstance(pair, dict):
        return False
    if (pair.get("sig") or "").strip() if isinstance(pair.get("sig"), str) else pair.get("sig"):
        return True
    dm = pair.get("date_mode")
    if mysql_store and mysql_store.is_composite_date_mode(dm):
        return bool((pair.get("date_iso") or "").strip())
    return bool((pair.get("date") or "").strip() if isinstance(pair.get("date"), str) else pair.get("date"))


def _merge_role_material_maps(base: dict, override: dict) -> dict:
    from sign_handlers.config import normalize_role_signer_map

    out: dict = {}
    for src in (base or {}, override or {}):
        if not isinstance(src, dict):
            continue
        for rk, rv in src.items():
            rid = str(rk or "").strip()
            if not rid or not isinstance(rv, dict):
                continue
            cur = dict(out.get(rid) or {})
            for k, v in rv.items():
                if k not in ("sig", "date", "date_mode", "date_iso"):
                    continue
                if v is None:
                    cur[k] = None
                elif isinstance(v, str):
                    if v.strip():
                        cur[k] = v.strip()
                else:
                    cur[k] = v
            if cur:
                out[rid] = cur
    return normalize_role_signer_map(out)


def _load_stroke_png_resolved(
    ref_id,
    kind: str,
    *,
    using_mysql: bool,
    mysql_store,
    local_get_item,
    inbox_root: str,
    sid: str,
    locale_hint: Optional[str] = None,
) -> Optional[bytes]:
    """加载 sig/date PNG：与批量签字、GET /stroke-items 同链路（item → signer → stroke_set）。"""
    iid = str(ref_id or "").strip()
    if not iid:
        return None
    k = (kind or "sig").strip().lower()
    if k not in ("sig", "date"):
        k = "sig"
    loc = (locale_hint or "").strip().lower()
    if loc not in ("zh", "en"):
        loc = ""

    if using_mysql and mysql_store:
        row = mysql_store.get_stroke_item_row(iid) or {}
        png = _coerce_png_bytes(row.get("png"))
        if png:
            return png
        signer_id = str(row.get("signer_id") or "").strip()
        item_loc = str(row.get("locale") or loc or "zh").strip().lower()
        if signer_id:
            try:
                alt = mysql_store.get_signer_stroke_png_resolved(
                    signer_id,
                    k,
                    item_loc if item_loc in ("zh", "en") else (loc or None),
                )
                png2 = _coerce_png_bytes(alt)
                if png2:
                    return png2
            except Exception:
                pass
            try:
                row2 = mysql_store.get_stroke_item_row_by_signer_kind(
                    signer_id,
                    item_loc if item_loc in ("zh", "en") else (loc or "zh"),
                    k,
                )
                png3 = _coerce_png_bytes((row2 or {}).get("png"))
                if png3:
                    return png3
            except Exception:
                pass
        try:
            alt_set = mysql_store.get_stroke_set_stroke_png_resolved(
                iid, k, loc or None
            )
            png4 = _coerce_png_bytes(alt_set)
            if png4:
                return png4
        except Exception:
            pass
        return None

    if local_get_item:
        return _coerce_png_bytes(local_get_item(inbox_root, sid, iid))
    return None


def _load_stroke_item_png(
    item_id,
    *,
    using_mysql: bool,
    mysql_store,
    local_get_item,
    inbox_root: str,
    sid: str,
) -> Optional[bytes]:
    return _load_stroke_png_resolved(
        item_id,
        "sig",
        using_mysql=using_mysql,
        mysql_store=mysql_store,
        local_get_item=local_get_item,
        inbox_root=inbox_root,
        sid=sid,
    )


def _resolve_pair_material_bytes(
    pair: dict,
    *,
    using_mysql: bool,
    mysql_store,
    local_get_item,
    inbox_root: str,
    sid: str,
) -> Tuple[Optional[bytes], Optional[bytes], Optional[str]]:
    """解析 role-map 条目为 (sig_png, date_png, error)。"""
    if not isinstance(pair, dict):
        return None, None, "角色素材未配置"
    sig_id = str(pair.get("sig") or "").strip() or None
    date_id = str(pair.get("date") or "").strip() or None
    dm = pair.get("date_mode")
    diso = str(pair.get("date_iso") or "").strip() or None

    sig_png = None
    date_png = None
    err: Optional[str] = None

    if sig_id:
        sig_png = _load_stroke_png_resolved(
            sig_id,
            "sig",
            using_mysql=using_mysql,
            mysql_store=mysql_store,
            local_get_item=local_get_item,
            inbox_root=inbox_root,
            sid=sid,
        )

    use_composite = (
        using_mysql
        and mysql_store
        and mysql_store.is_composite_date_mode(dm)
        and bool(diso)
    )
    if use_composite:
        if not sig_id:
            err = "拼接日期需绑定签名并填写日历日期"
        elif not sig_png:
            err = "签名素材 PNG 不可用"
        else:
            srow = mysql_store.get_stroke_item_row(sig_id) if sig_id else None
            sid0 = str((srow or {}).get("signer_id") or "").strip()
            if not sid0:
                err = "无法解析签署人（拼接日期）"
            else:
                try:
                    lay = mysql_store.composite_mode_to_layout(dm)
                    date_png, _ = mysql_store.compose_date_piece_png(sid0, diso, lay)
                    date_png = _coerce_png_bytes(date_png)
                    if not date_png:
                        err = "日期拼接结果为空"
                except Exception as e:
                    err = f"日期拼接失败：{e}"
    elif date_id:
        date_png = _load_stroke_png_resolved(
            date_id,
            "date",
            using_mysql=using_mysql,
            mysql_store=mysql_store,
            local_get_item=local_get_item,
            inbox_root=inbox_root,
            sid=sid,
        )

    return sig_png, date_png, err


def _preprocess_role_pair(
    sig_png: Optional[bytes],
    date_png: Optional[bytes],
    *,
    ext: str,
) -> Tuple[Optional[bytes], Optional[bytes]]:
    ext_l = (ext or "").lower()
    if sig_png and date_png:
        if ext_l in (".xlsx", ".xls"):
            return prepare_signature_date_pair_for_word(
                sig_png, date_png, equalize_ink_scale=False
            )
        return prepare_signature_date_pair_for_word(sig_png, date_png)
    sig = prepare_png_for_word(sig_png) if sig_png else None
    dt = prepare_png_for_word(date_png) if date_png else None
    if sig is None and sig_png:
        sig = sig_png
    if dt is None and date_png:
        dt = date_png
    return sig, dt


def _resolve_xlsx_target_cell(wb, role_plan: dict, kind: str):
    from sign_handlers.sign_xlsx import _is_emptyish, _xlsx_row_skip_for_signoff

    if kind == "sig":
        spec = role_plan.get("xlsx_name_cell") or role_plan.get("name_cell")
    else:
        spec = role_plan.get("xlsx_date_cell") or role_plan.get("date_cell")
    if isinstance(spec, dict):
        sno = int(spec.get("sheet") or spec.get("table") or 1)
        rno = int(spec.get("row") or 0)
        cno = int(spec.get("col") or 0)
        if rno >= 1 and cno >= 1:
            if sno < 1 or sno > len(wb.worksheets):
                ws = wb.worksheets[0]
            else:
                ws = wb.worksheets[sno - 1]
            if _xlsx_row_skip_for_signoff(ws, rno):
                pass
            else:
                right_c = cno + 1
                if right_c <= int(ws.max_column or 1):
                    right = ws.cell(row=rno, column=right_c)
                    if _is_emptyish(right.value):
                        return ws, right, 0
                return ws, ws.cell(row=rno, column=cno), 0

    hints = role_plan.get("source_hints") if isinstance(role_plan.get("source_hints"), list) else []
    from sign_handlers.sign_xlsx import _parse_xlsx_row_hint

    for hint in hints:
        parsed = _parse_xlsx_row_hint(str(hint or ""))
        if not parsed:
            continue
        sno, rno = parsed
        ws_list = list(wb.worksheets) if sno is None else [wb.worksheets[sno - 1]]
        for ws in ws_list:
            if rno < 1 or rno > int(ws.max_row or 1):
                continue
            if _xlsx_row_skip_for_signoff(ws, rno):
                continue
            max_c = min(int(ws.max_column or 1), 128)
            for cno in range(1, max_c + 1):
                right = ws.cell(row=rno, column=cno + 1) if cno + 1 <= max_c else None
                if right is not None and _is_emptyish(right.value):
                    return ws, right, 0
                if _is_emptyish(ws.cell(row=rno, column=cno).value):
                    return ws, ws.cell(row=rno, column=cno), 0
    return None, None, 0


def _default_xlsx_insert_box(wb) -> Tuple[int, int]:
    from sign_handlers.sign_xlsx import (
        _EXCEL_ABS_MAX_IMG_WIDTH_PX,
        _excel_content_max_width_px,
        _iter_xlsx_cover_worksheets,
        _merged_span_box_px,
        _xlsx_row_has_cover_signoff_labels,
        _xlsx_row_skip_for_signoff,
    )

    ws_candidates = _iter_xlsx_cover_worksheets(wb)
    if not ws_candidates:
        ws_candidates = [wb.active if wb.active is not None else wb.worksheets[0]]
    for ws in ws_candidates:
        max_c = min(int(ws.max_column or 1), 128)
        for rr in range(1, min(int(ws.max_row or 1), 64) + 1):
            if _xlsx_row_has_cover_signoff_labels(ws, rr, max_c=max_c):
                cell = ws.cell(row=rr, column=min(4, max_c))
                box_w, box_h = _merged_span_box_px(ws, cell)
                max_w2 = min(int(_EXCEL_ABS_MAX_IMG_WIDTH_PX), _excel_content_max_width_px(int(box_w)))
                max_h2 = max(14, int(box_h) - 4)
                return max(24, max_w2), max_h2
    ws = ws_candidates[0]
    r_end = int(ws.max_row or 1)
    target_r = r_end
    for rr in range(r_end, 0, -1):
        if not _xlsx_row_skip_for_signoff(ws, rr):
            target_r = rr
            break
    cell = ws.cell(row=target_r, column=min(3, int(ws.max_column or 1) or 1))
    box_w, box_h = _merged_span_box_px(ws, cell)
    max_w2 = min(int(_EXCEL_ABS_MAX_IMG_WIDTH_PX), _excel_content_max_width_px(int(box_w)))
    max_h2 = max(14, int(box_h) - 4)
    return max(24, max_w2), max_h2


def _scale_for_xlsx_insert(
    png_bytes: bytes,
    file_bytes: bytes,
    role_id: str,
    kind: str,
    placement_plan: Optional[dict],
    *,
    wb=None,
) -> bytes:
    from openpyxl import load_workbook

    from sign_handlers.sign_xlsx import compute_xlsx_insert_pixel_size

    close_wb = False
    if wb is None:
        bio = io.BytesIO(file_bytes)
        wb = load_workbook(bio, data_only=True)
        close_wb = True
    try:
        rp = (placement_plan or {}).get(str(role_id)) if isinstance(placement_plan, dict) else {}
        rp = rp if isinstance(rp, dict) else {}
        ws, cell, inline_x = _resolve_xlsx_target_cell(wb, rp, kind)
        if ws is not None and cell is not None:
            tw, th = compute_xlsx_insert_pixel_size(
                png_bytes, ws, cell, inline_offset_px=int(inline_x or 0)
            )
        else:
            tw, th = _default_xlsx_insert_box(wb)
        scaled = scale_png_to_pixel_box(png_bytes, tw, th)
        return scaled or png_bytes
    finally:
        if close_wb:
            try:
                wb.close()
            except Exception:
                pass


def _finalize_export_png(
    png_bytes: Optional[bytes],
    *,
    ext: str,
    file_bytes: bytes,
    role_id: str,
    kind: str,
    placement_plan: Optional[dict],
    wb=None,
) -> Optional[bytes]:
    if not png_bytes:
        return None
    ext_l = (ext or "").lower()
    if ext_l in (".docx", ".doc"):
        return scale_png_for_docx_insert(png_bytes) or png_bytes
    if ext_l in (".xlsx", ".xls"):
        return _scale_for_xlsx_insert(
            png_bytes, file_bytes, role_id, kind, placement_plan, wb=wb
        )
    return png_bytes


def _pair_with_doc_date(pair: dict, doc_date_fallback: str) -> dict:
    p = dict(pair) if isinstance(pair, dict) else {}
    if doc_date_fallback and not (p.get("date_iso") or "").strip():
        p["date_iso"] = str(doc_date_fallback).strip()[:10]
    return p


def export_role_material_png(
    *,
    file_bytes: bytes,
    ext: str,
    role_id: str,
    kind: str,
    pair: dict,
    placement_plan: Optional[dict] = None,
    using_mysql: bool,
    mysql_store=None,
    local_get_item=None,
    inbox_root: str = "",
    sid: str = "",
    wb=None,
) -> Tuple[Optional[bytes], Optional[str]]:
    """
    返回 (png_bytes, error)。
    kind: sig | date
    """
    kind_l = str(kind or "").strip().lower()
    if kind_l not in ("sig", "date"):
        return None, "kind 须为 sig 或 date"

    sig_raw, date_raw, err = _resolve_pair_material_bytes(
        pair,
        using_mysql=using_mysql,
        mysql_store=mysql_store,
        local_get_item=local_get_item,
        inbox_root=inbox_root,
        sid=sid,
    )
    if err and kind_l == "date" and not date_raw:
        return None, err

    sig_prep, date_prep = _preprocess_role_pair(sig_raw, date_raw, ext=ext)
    png = sig_prep if kind_l == "sig" else date_prep
    if not png:
        if kind_l == "sig":
            return None, "签名素材未选择或不可用"
        return None, err or "日期素材未选择或不可用"

    out = _finalize_export_png(
        png,
        ext=ext,
        file_bytes=file_bytes,
        role_id=role_id,
        kind=kind_l,
        placement_plan=placement_plan,
        wb=wb,
    )
    return out, None


def export_role_materials_zip(
    *,
    file_bytes: bytes,
    ext: str,
    role_map: dict,
    placement_plan: Optional[dict] = None,
    using_mysql: bool,
    mysql_store=None,
    local_get_item=None,
    inbox_root: str = "",
    sid: str = "",
    doc_date_fallback: str = "",
) -> Tuple[Optional[bytes], Optional[str], int]:
    """批量导出已有素材为 ZIP（仅打开 Excel 一次，避免多次请求卡顿）。"""
    from openpyxl import load_workbook

    role_map = role_map if isinstance(role_map, dict) else {}
    wb = None
    ext_l = (ext or "").lower()
    if ext_l in (".xlsx", ".xls"):
        try:
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            return None, f"读取 Excel 失败：{e}", 0

    zbuf = io.BytesIO()
    count = 0
    misses: list[str] = []
    try:
        with zipfile.ZipFile(zbuf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for role_id in _EXPORT_ROLE_ORDER:
                pair = _pair_with_doc_date(role_map.get(role_id) or {}, doc_date_fallback)
                if not _role_map_entry_nonempty(pair, mysql_store=mysql_store):
                    continue
                sig_id = str(pair.get("sig") or "").strip() or None
                date_id = str(pair.get("date") or "").strip() or None
                dm = pair.get("date_mode")
                diso = str(pair.get("date_iso") or "").strip() or None
                sig_raw, date_raw, err = _resolve_pair_material_bytes(
                    pair,
                    using_mysql=using_mysql,
                    mysql_store=mysql_store,
                    local_get_item=local_get_item,
                    inbox_root=inbox_root,
                    sid=sid,
                )
                sig_prep, date_prep = _preprocess_role_pair(sig_raw, date_raw, ext=ext)
                for kind in _EXPORT_KIND_ORDER:
                    tag = MATERIAL_SHORT_TAGS.get((role_id, kind), f"{role_id}_{kind}")
                    if kind == "sig":
                        if not sig_id:
                            continue
                        if not sig_raw:
                            misses.append(f"{tag}：签名素材未绑定或 PNG 不可用")
                            continue
                    else:
                        if not date_raw:
                            if mysql_store and mysql_store.is_composite_date_mode(dm):
                                if not diso:
                                    misses.append(f"{tag}：缺文档体现日期")
                                elif err:
                                    misses.append(f"{tag}：{err}")
                                else:
                                    misses.append(f"{tag}：日期素材未绑定或 PNG 不可用")
                            elif not date_id:
                                continue
                            else:
                                misses.append(f"{tag}：日期素材未绑定或 PNG 不可用")
                            continue
                    src = sig_prep if kind == "sig" else date_prep
                    if not src:
                        misses.append(f"{tag}：PNG 预处理失败")
                        continue
                    png = _finalize_export_png(
                        src,
                        ext=ext,
                        file_bytes=file_bytes,
                        role_id=role_id,
                        kind=kind,
                        placement_plan=placement_plan,
                        wb=wb,
                    )
                    if not png:
                        misses.append(f"{tag}：缩放导出失败")
                        continue
                    zf.writestr(f"{tag}.png", png)
                    count += 1
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception:
                pass

    if count <= 0:
        hint = "；".join(misses[:8]) if misses else "role-map 中无可用签名/日期绑定"
        return None, f"没有可导出的签名/日期素材（{hint}）", 0
    return zbuf.getvalue(), None, count
