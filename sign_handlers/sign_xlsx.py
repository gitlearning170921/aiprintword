# -*- coding: utf-8 -*-
"""Excel .xlsx：在关键词单元格相邻空白格插入签名/日期 PNG。"""
from __future__ import annotations

import io
import os
import re
from typing import Optional

from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils.units import pixels_to_EMU
from openpyxl import load_workbook

from sign_handlers.config import (
    ROLE_ID_TO_KEYWORD,
    is_replaceable_prefilled_slot_text,
    role_keywords_for_apply,
)
from sign_handlers.label_match import (
    cell_has_label_inline_reservation,
    cell_has_role_keyword,
    cell_has_signoff_inline_reservation,
    cell_inline_insert_offset_px,
    cell_is_bare_role_column_header,
    cell_is_role_signoff_label_slot,
    cell_looks_like_signoff_date_label,
    cell_text_matches_keyword,
    xlsx_cell_has_leading_role_keyword,
)
from sign_handlers.png_word_compat import (
    prepare_png_for_word,
    prepare_signature_date_pair_for_word,
)

# Excel：按合并区域/单元格真实宽高适配；优先吃满行高（等比缩放、不变形），
# 不设过小全局宽度上限，避免「行高还够但宽度先被 420 卡住」导致图偏小发虚。
# 绝对宽度仅作防爆栅格，适度允许放大以填满行（上限避免糊成一片）。
_EXCEL_ABS_MAX_IMG_WIDTH_PX = 1600
_EXCEL_MAX_UPSCALE = 2.5
_DATE_LABEL_KEYWORDS = ("日期", "Date")
_GAP_PX = 6


def _planned_keywords_for_role(
    role_id: str,
    placement_plan: dict | None,
) -> list[str]:
    if not isinstance(placement_plan, dict):
        return []
    role_plan = placement_plan.get(str(role_id))
    if not isinstance(role_plan, dict):
        return []
    kws = role_plan.get("keywords")
    if not isinstance(kws, list):
        return []
    out: list[str] = []
    seen: set[str] = set()
    for item in kws:
        s = str(item or "").strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    out.sort(key=len, reverse=True)
    return out


def _planned_source_hints_for_role(
    role_id: str,
    placement_plan: dict | None,
) -> list[str]:
    if not isinstance(placement_plan, dict):
        return []
    role_plan = placement_plan.get(str(role_id))
    if not isinstance(role_plan, dict):
        return []
    arr = role_plan.get("source_hints")
    if not isinstance(arr, list):
        return []
    out: list[str] = []
    seen: set[str] = set()
    for item in arr:
        s = str(item or "").strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def _planned_layout_types_for_role(
    role_id: str,
    placement_plan: dict | None,
) -> list[str]:
    if not isinstance(placement_plan, dict):
        return []
    role_plan = placement_plan.get(str(role_id))
    if not isinstance(role_plan, dict):
        return []
    arr = role_plan.get("layout_types")
    if not isinstance(arr, list):
        return []
    out: list[str] = []
    seen: set[str] = set()
    for item in arr:
        s = str(item or "").strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def _parse_xlsx_row_hint(source_hint: str) -> tuple[int | None, int] | None:
    s = str(source_hint or "").strip()
    m = re.match(r"^\s*sheet(\d+)\.row(\d+)\s*$", s, re.IGNORECASE)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^\s*table(\d+)\.row(\d+)\s*$", s, re.IGNORECASE)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^\s*row(\d+)\s*$", s, re.IGNORECASE)
    if m:
        return None, int(m.group(1))
    return None


def _col_width_px(ws, col_idx: int) -> int:
    """Excel 列宽到像素的近似换算（足够用于锚点/缩放）。"""
    try:
        letter = get_column_letter(int(col_idx))
        w = ws.column_dimensions[letter].width
        if w is None:
            w = ws.sheet_format.defaultColWidth
        if w is None:
            w = 8.43
        # 经验公式：1 字符宽约 7px + 5px padding
        return max(24, int(float(w) * 7.0 + 5.0))
    except Exception:
        return 64


def _row_height_px(ws, row_idx: int) -> int:
    """行高（pt）到像素近似换算。"""
    try:
        h = ws.row_dimensions[int(row_idx)].height
        if h is None:
            h = ws.sheet_format.defaultRowHeight
        if h is None:
            h = 15.0
        return max(18, int(float(h) * 96.0 / 72.0))
    except Exception:
        return 20


def _merged_span_box_px(ws, cell) -> tuple[int, int]:
    """
    合并单元格：用合并区域累计列宽、累计行高，避免图片只按左上角单列宽缩放导致「撑出格子」。
    非合并：退化为单列宽 + 单行高。
    """
    try:
        merged = getattr(ws, "merged_cells", None)
        if merged is None:
            return _col_width_px(ws, cell.column), _row_height_px(ws, cell.row)
        for rng in merged.ranges:
            if cell.coordinate in rng:
                wpx = 0
                for col in range(int(rng.min_col), int(rng.max_col) + 1):
                    wpx += _col_width_px(ws, col)
                hpx = 0
                for row in range(int(rng.min_row), int(rng.max_row) + 1):
                    hpx += _row_height_px(ws, row)
                return max(24, wpx), max(18, hpx)
    except Exception:
        pass
    return _col_width_px(ws, cell.column), _row_height_px(ws, cell.row)


def _excel_content_max_width_px(box_w: int) -> int:
    """单元格可用内容宽度（像素近似），并加绝对上限防爆。"""
    return min(max(24, int(box_w) - 4), _EXCEL_ABS_MAX_IMG_WIDTH_PX)


def _fit_to_box_excel(w: int, h: int, max_w: int, max_h: int) -> tuple[int, int]:
    """
    等比装入 max_w×max_h，不变形；不强制「只缩小不放大」，便于在小 PNG 时吃满行高更清晰。
    缩放系数取 min(max_w/w, max_h/h)，行高较紧时自然由行高主导。
    """
    if w <= 0 or h <= 0:
        return max(1, max_w), max(1, max_h)
    s = min(max_w / float(w), max_h / float(h))
    s = min(s, _EXCEL_MAX_UPSCALE)
    return max(1, int(round(w * s))), max(1, int(round(h * s)))


def _find_date_cell_same_row(ws, r: int, author_col: int):
    """同行查找「Date / 日期」标签格与其右侧空白格。"""
    max_col = min(ws.max_column or 0, 64)
    if max_col < author_col + 2:
        return None, None
    for col in range(author_col + 2, max_col + 1):
        cl = ws.cell(row=r, column=col)
        for dkw in _DATE_LABEL_KEYWORDS:
            if cell_text_matches_keyword(cl.value, dkw):
                dc = ws.cell(row=r, column=col + 1)
                if _is_emptyish(dc.value):
                    return cl, dc
                return cl, None
    return None, None


def _is_emptyish(v) -> bool:
    if v is None:
        return True
    s = str(v).strip()
    if not s:
        return True
    if all(c in " _-—–\u2014\u2015\u2500\u3000.·" for c in s):
        return True
    return len(s) < 2


def _is_slot_target(v) -> bool:
    """签名/日期可落位目标：空白占位，或可替换的电脑输入值。"""
    return _is_emptyish(v) or is_replaceable_prefilled_slot_text(str(v or ""))


def _has_date_context_nearby(ws, r: int, c: int, val) -> bool:
    """
    限制“可替换已有值”只在签批语境生效，避免把普通业务表格列（如风险分析里的“审核”）误判为签字位。
    """
    if _find_date_cell_same_row(ws, r, c)[0] is not None:
        return True
    max_c = int(ws.max_column or 1)
    for dc in (c + 1, c + 2, c + 3):
        if dc > max_c:
            continue
        if cell_text_matches_keyword(ws.cell(row=r, column=dc).value, "日期") or cell_text_matches_keyword(
            ws.cell(row=r, column=dc).value, "Date"
        ):
            return True
    if r + 1 <= int(ws.max_row or 1):
        for dc in (c, c + 1, c + 2):
            if dc > max_c:
                continue
            if cell_text_matches_keyword(ws.cell(row=r + 1, column=dc).value, "日期") or cell_text_matches_keyword(
                ws.cell(row=r + 1, column=dc).value, "Date"
            ):
                return True
    return False


def _add_png(
    ws,
    png_bytes: bytes,
    anchor: str,
    layout_cell=None,
    max_w: int = _EXCEL_ABS_MAX_IMG_WIDTH_PX,
) -> None:
    img = XLImage(io.BytesIO(png_bytes))
    w = int(getattr(img, "width", None) or max_w)
    h = int(getattr(img, "height", None) or int(max_w * 0.35))
    lc = layout_cell
    if lc is None:
        try:
            r, c = coordinate_to_tuple(anchor)
            lc = ws.cell(row=r, column=c)
        except Exception:
            lc = None
    try:
        box_w, box_h = _merged_span_box_px(ws, lc) if lc is not None else (int(max_w), int(max_w * 0.55))
    except Exception:
        box_w, box_h = int(max_w), int(max_w * 0.55)
    max_w2 = min(int(max_w), _excel_content_max_width_px(int(box_w)))
    max_h2 = max(14, int(box_h) - 4)
    w2, h2 = _fit_to_box_excel(w, h, max_w2, max_h2)
    img.width = w2
    img.height = h2
    img.anchor = anchor
    ws.add_image(img)


def _merged_top_left_coordinate(ws, cell) -> str:
    """
    若 cell 位于合并单元格内，返回该合并区域左上角坐标；否则返回 cell.coordinate。
    模板里常见「标题+点线留白」做成合并格，锚到右邻格会导致图片看起来不贴标题。
    """
    try:
        merged = getattr(ws, "merged_cells", None)
        if merged is None:
            return cell.coordinate
        for rng in merged.ranges:
            if cell.coordinate in rng:
                return f"{get_column_letter(rng.min_col)}{rng.min_row}"
    except Exception:
        pass
    return cell.coordinate


def _add_png_after_label_in_cell(
    ws, png_bytes: bytes, cell, offset_x_px: int, max_w: int = _EXCEL_ABS_MAX_IMG_WIDTH_PX
) -> None:
    """把图片锚到当前标签单元格内偏移位置，避免落到占位点线末端，并按格宽高自适应缩放。"""
    img = XLImage(io.BytesIO(png_bytes))
    w = int(getattr(img, "width", None) or max_w)
    h = int(getattr(img, "height", None) or int(max_w * 0.35))
    cell_w, cell_h = _merged_span_box_px(ws, cell)
    avail_w = max(28, cell_w - int(offset_x_px) - 4)
    max_w2 = min(int(max_w), int(avail_w), _excel_content_max_width_px(cell_w))
    max_h2 = max(14, int(cell_h) - 4)
    w2, h2 = _fit_to_box_excel(w, h, max_w2, max_h2)
    img.width = w2
    img.height = h2
    marker = AnchorMarker(
        col=cell.column - 1,
        row=cell.row - 1,
        colOff=pixels_to_EMU(offset_x_px),
        rowOff=pixels_to_EMU(1),
    )
    ext = XDRPositiveSize2D(pixels_to_EMU(int(img.width)), pixels_to_EMU(int(img.height)))
    img.anchor = OneCellAnchor(_from=marker, ext=ext)
    ws.add_image(img)


def _add_sig_and_date_after_label_in_cell(
    ws,
    sig_png: bytes,
    date_png: bytes,
    cell,
    offset_x_px: int,
    max_w: int = _EXCEL_ABS_MAX_IMG_WIDTH_PX,
) -> None:
    """同一单元格内：标题后先签名紧接日期（两个图片各自锚在不同 colOff）。"""
    cell_w, cell_h = _merged_span_box_px(ws, cell)
    avail_w = max(40, cell_w - int(offset_x_px) - 4)
    max_h2 = max(14, int(cell_h) - 4)

    sig_img = XLImage(io.BytesIO(sig_png))
    s_w = int(getattr(sig_img, "width", None) or max_w)
    s_h = int(getattr(sig_img, "height", None) or int(max_w * 0.35))
    date_img = XLImage(io.BytesIO(date_png))
    d_w = int(getattr(date_img, "width", None) or max_w)
    d_h = int(getattr(date_img, "height", None) or int(max_w * 0.35))

    # 先给签名一个上限（优先留位置给日期）
    sig_box_w = max(24, min(int(max_w), int(avail_w * 0.62), _excel_content_max_width_px(cell_w)))
    s_w2, s_h2 = _fit_to_box_excel(s_w, s_h, sig_box_w, max_h2)
    remain = max(24, avail_w - s_w2 - _GAP_PX)
    d_box_w = max(24, min(int(max_w), int(remain), _excel_content_max_width_px(cell_w)))
    d_w2, d_h2 = _fit_to_box_excel(d_w, d_h, d_box_w, max_h2)

    sig_img.width = s_w2
    sig_img.height = s_h2
    date_img.width = d_w2
    date_img.height = d_h2

    m1 = AnchorMarker(
        col=cell.column - 1,
        row=cell.row - 1,
        colOff=pixels_to_EMU(int(offset_x_px)),
        rowOff=pixels_to_EMU(1),
    )
    e1 = XDRPositiveSize2D(pixels_to_EMU(int(sig_img.width)), pixels_to_EMU(int(sig_img.height)))
    sig_img.anchor = OneCellAnchor(_from=m1, ext=e1)
    ws.add_image(sig_img)

    m2 = AnchorMarker(
        col=cell.column - 1,
        row=cell.row - 1,
        colOff=pixels_to_EMU(int(offset_x_px) + int(sig_img.width) + _GAP_PX),
        rowOff=pixels_to_EMU(1),
    )
    e2 = XDRPositiveSize2D(pixels_to_EMU(int(date_img.width)), pixels_to_EMU(int(date_img.height)))
    date_img.anchor = OneCellAnchor(_from=m2, ext=e2)
    ws.add_image(date_img)


def _is_xlsx_label_blank_date_blank_row(
    ws, r: int, label_c: int, kw: str, max_c: int
) -> bool:
    """Excel 四列：角色 | 姓名空白 | 日期 | 日期空白。"""
    if label_c + 3 > max_c:
        return False
    lv = ws.cell(row=r, column=label_c).value
    if not cell_has_role_keyword(str(lv or ""), kw):
        return False
    if not _is_slot_target(ws.cell(row=r, column=label_c + 1).value):
        return False
    dv = ws.cell(row=r, column=label_c + 2).value
    if not (
        cell_text_matches_keyword(dv, "日期")
        or cell_text_matches_keyword(dv, "Date")
        or cell_looks_like_signoff_date_label(dv)
    ):
        return False
    return _is_slot_target(ws.cell(row=r, column=label_c + 3).value)


def _try_xlsx_four_column_row(
    ws, r: int, label_c: int, kw: str, sig, dt, max_r: int
) -> bool:
    max_c = min(int(ws.max_column or 1), 128)
    if not _is_xlsx_label_blank_date_blank_row(ws, r, label_c, kw, max_c):
        return False
    placed = False
    sig_cell = ws.cell(row=r, column=label_c + 1)
    date_hdr = ws.cell(row=r, column=label_c + 2)
    date_cell = ws.cell(row=r, column=label_c + 3)
    if sig and _is_slot_target(sig_cell.value):
        sig_cell.value = None
        _add_png(ws, sig, _merged_top_left_coordinate(ws, sig_cell), sig_cell)
        placed = True
    if dt:
        if _is_slot_target(date_cell.value):
            date_cell.value = None
            _add_png(ws, dt, _merged_top_left_coordinate(ws, date_cell), date_cell)
            placed = True
        else:
            d_inline = cell_inline_insert_offset_px(date_hdr.value, "Date")
            if d_inline is None:
                d_inline = cell_inline_insert_offset_px(date_hdr.value, "日期")
            if d_inline is not None:
                _add_png_after_label_in_cell(ws, dt, date_hdr, d_inline)
                placed = True
            elif _is_slot_target(ws.cell(row=r, column=label_c + 4).value):
                dc2 = ws.cell(row=r, column=label_c + 4)
                dc2.value = None
                _add_png(ws, dt, _merged_top_left_coordinate(ws, dc2), dc2)
                placed = True
    if sig and not dt:
        return placed
    if dt:
        return placed
    return placed


def _try_xlsx_role_layout_cells(
    wb,
    role_id: str,
    placement_plan: dict | None,
    keywords: list,
    sig,
    dt,
) -> bool:
    if not isinstance(placement_plan, dict):
        return False
    rp = placement_plan.get(str(role_id))
    if not isinstance(rp, dict):
        return False
    name_cell = rp.get("xlsx_name_cell") or rp.get("name_cell")
    if not isinstance(name_cell, dict):
        return False
    want_date = bool(rp.get("date_slot", True))
    dt_use = dt if want_date else None
    sno = int(name_cell.get("sheet") or name_cell.get("table") or 0)
    rno = int(name_cell.get("row") or 0)
    cno = int(name_cell.get("col") or 0)
    if rno < 1 or cno < 1:
        return False
    if sno < 1 or sno > len(wb.worksheets):
        ws_list = list(wb.worksheets)
    else:
        ws_list = [wb.worksheets[sno - 1]]
    max_r = 0
    for ws in ws_list:
        max_r = int(ws.max_row or 1)
        for kw in keywords:
            for c in range(max(1, cno - 1), min(int(ws.max_column or 1), 128) + 1):
                if _try_xlsx_four_column_row(ws, rno, c, kw, sig, dt_use, max_r):
                    return True
                if _try_place_at_keyword_cell(ws, rno, c, kw, sig, dt_use, max_r):
                    return True
            for c in range(1, min(int(ws.max_column or 1), 128) + 1):
                if _try_xlsx_four_column_row(ws, rno, c, kw, sig, dt_use, max_r):
                    return True
    return False


def _try_place_at_keyword_cell(ws, r: int, c: int, kw: str, sig, dt, max_r: int) -> bool:
    max_c = min(int(ws.max_column or 1), 128)
    for label_c in range(max(1, c - 1), min(c + 2, max_c + 1)):
        if _try_xlsx_four_column_row(ws, r, label_c, kw, sig, dt, max_r):
            return True
    cell = ws.cell(row=r, column=c)
    val = cell.value
    ct = str(val).strip() if val is not None else ""
    if not (cell_text_matches_keyword(val, kw) or xlsx_cell_has_leading_role_keyword(val, kw)):
        return False
    if cell_is_bare_role_column_header(val, kw):
        if _try_xlsx_four_column_row(ws, r, c, kw, sig, dt, max_r):
            return True
        return False
    has_blank_right = c + 1 <= int(ws.max_column or 1) and _is_emptyish(
        ws.cell(row=r, column=c + 1).value
    )
    has_blank_below = r + 1 <= max_r and _is_emptyish(ws.cell(row=r + 1, column=c).value)
    signoff_slot = cell_is_role_signoff_label_slot(val, kw)
    inline_slot = cell_has_label_inline_reservation(val, kw)
    date_ctx = _has_date_context_nearby(ws, r, c, val)
    if not (signoff_slot or inline_slot or ((has_blank_right or has_blank_below) and date_ctx)):
        return False
    sig_cell = ws.cell(row=r, column=c + 1)
    inline_x = cell_inline_insert_offset_px(val, kw) if signoff_slot else None
    if sig and (not _is_slot_target(sig_cell.value)):
        if inline_x is None:
            return False
    date_label_cell, date_cell = _find_date_cell_same_row(ws, r, c)
    if date_cell is None:
        c2 = ws.cell(row=r, column=c + 2)
        if _is_slot_target(c2.value):
            date_cell = c2
        else:
            d1 = ws.cell(row=r + 1, column=c + 1)
            if _is_slot_target(d1.value):
                date_cell = d1
            else:
                d0 = ws.cell(row=r + 1, column=c)
                if _is_slot_target(d0.value):
                    date_cell = d0
    placed_any = False
    if inline_x is not None and sig and dt:
        _add_sig_and_date_after_label_in_cell(ws, sig, dt, cell, inline_x)
        placed_any = True
    elif inline_x is not None and sig:
        _add_png_after_label_in_cell(ws, sig, cell, inline_x)
        placed_any = True
    if sig and not placed_any:
        if _is_slot_target(sig_cell.value):
            sig_cell.value = None
            _add_png(
                ws,
                sig,
                _merged_top_left_coordinate(ws, sig_cell),
                sig_cell,
            )
            placed_any = True
        if not placed_any and has_blank_below:
            below = ws.cell(row=r + 1, column=c)
            below.value = None
            _add_png(
                ws,
                sig,
                _merged_top_left_coordinate(ws, below),
                below,
            )
            placed_any = True
        if not placed_any:
            sig_inline_x = cell_inline_insert_offset_px(cell.value, kw)
            if sig_inline_x is not None:
                if dt and cell_has_signoff_inline_reservation(val, kw):
                    _add_sig_and_date_after_label_in_cell(
                        ws, sig, dt, cell, sig_inline_x
                    )
                else:
                    _add_png_after_label_in_cell(
                        ws, sig, cell, sig_inline_x
                    )
                placed_any = True
    if dt and not placed_any:
        if signoff_slot:
            max_c = int(ws.max_column or 1)
            for dc in range(c + 2, min(c + 8, max_c + 1)):
                dcell = ws.cell(row=r, column=dc)
                dv = dcell.value
                if (
                    cell_looks_like_signoff_date_label(dv)
                    and not cell_has_role_keyword(dv, kw)
                ):
                    break
                if _is_slot_target(dv):
                    dcell.value = None
                    _add_png(
                        ws,
                        dt,
                        _merged_top_left_coordinate(ws, dcell),
                        dcell,
                    )
                    placed_any = True
                    break
        if (
            not placed_any
            and signoff_slot
            and cell_has_signoff_inline_reservation(val, kw)
        ):
            inline_dt_x = cell_inline_insert_offset_px(val, kw)
            if inline_dt_x is not None:
                if sig:
                    _add_sig_and_date_after_label_in_cell(
                        ws, sig, dt, cell, inline_dt_x
                    )
                else:
                    _add_png_after_label_in_cell(
                        ws, dt, cell, inline_dt_x
                    )
                placed_any = True
        if (
            not placed_any
            and not signoff_slot
            and date_cell is not None
            and _is_slot_target(date_cell.value)
        ):
            date_cell.value = None
            _add_png(ws, dt, _merged_top_left_coordinate(ws, date_cell), date_cell)
            placed_any = True
        elif not placed_any:
            date_inline_x = None
            if date_label_cell is not None:
                date_inline_x = cell_inline_insert_offset_px(date_label_cell.value, "Date")
                if date_inline_x is None:
                    date_inline_x = cell_inline_insert_offset_px(date_label_cell.value, "日期")
            if date_label_cell is not None and date_inline_x is not None:
                _add_png_after_label_in_cell(ws, dt, date_label_cell, date_inline_x)
                placed_any = True
            elif date_cell is not None:
                date_cell.value = None
                _add_png(ws, dt, _merged_top_left_coordinate(ws, date_cell), date_cell)
                placed_any = True
            elif (not _is_slot_target(sig_cell.value)) and sig:
                dnext = ws.cell(row=r, column=c + 2)
                if _is_slot_target(dnext.value):
                    dnext.value = None
                    _add_png(
                        ws,
                        dt,
                        _merged_top_left_coordinate(ws, dnext),
                        dnext,
                    )
                    placed_any = True
            if not placed_any:
                below = f"{get_column_letter(c + 1)}{r + 1}"
                below_cell = ws.cell(row=r + 1, column=c + 1)
                _add_png(ws, dt, below, below_cell)
                placed_any = True
    return placed_any


def sign_xlsx(
    path: str,
    role_to_signature_png: dict,
    role_to_date_png: dict,
    out_path: Optional[str] = None,
    placement_plan: dict | None = None,
    placement_result: dict | None = None,
) -> str:
    path = os.path.abspath(path)
    if out_path is None:
        base, ext = os.path.splitext(path)
        out_path = f"{base}_signed{ext}"

    wb = load_workbook(path)
    for role_id in ROLE_ID_TO_KEYWORD:
        sb = role_to_signature_png.get(role_id)
        db = role_to_date_png.get(role_id)
        if sb and db:
            sig, dt = prepare_signature_date_pair_for_word(
                sb, db, equalize_ink_scale=False
            )
        else:
            sig = prepare_png_for_word(sb) if sb else None
            dt = prepare_png_for_word(db) if db else None
            if sig is None and sb:
                sig = sb
            if dt is None and db:
                dt = db
        want_date = True
        if isinstance(placement_plan, dict):
            rp0 = placement_plan.get(str(role_id))
            if isinstance(rp0, dict) and "date_slot" in rp0:
                want_date = bool(rp0.get("date_slot"))
        dt_use = dt if want_date else None
        if not sig and not dt_use:
            if isinstance(placement_result, dict):
                placement_result[str(role_id)] = {
                    "applied_sig": False,
                    "applied_date": False,
                    "placed": False,
                    "placed_by": "skip_no_material",
                    "keywords": [],
                }
            continue
        placed = False
        placed_by = ""
        fallback_kws = sorted(role_keywords_for_apply(role_id), key=lambda x: len(x), reverse=True)
        planned_kws = _planned_keywords_for_role(role_id, placement_plan)
        planned_hints = _planned_source_hints_for_role(role_id, placement_plan)
        planned_layout_types = _planned_layout_types_for_role(role_id, placement_plan)
        kws = []
        seen_kw: set[str] = set()
        for kw in planned_kws + fallback_kws:
            if kw in seen_kw:
                continue
            seen_kw.add(kw)
            kws.append(kw)
        if isinstance(placement_result, dict):
            placement_result[str(role_id)] = {
                "applied_sig": bool(sig),
                "applied_date": bool(dt),
                "placed": False,
                "placed_by": "",
                "keywords": kws[:],
                "source_hints": planned_hints[:],
                "layout_types": planned_layout_types[:],
            }
        if (not placed) and placement_plan:
            placed = _try_xlsx_role_layout_cells(
                wb, role_id, placement_plan, kws, sig, dt_use
            )
            if placed:
                placed_by = "layout_cells"
        if planned_hints and kws:
            for hint in planned_hints:
                parsed = _parse_xlsx_row_hint(hint)
                if not parsed:
                    continue
                sheet_no, row_no = parsed
                if row_no <= 0:
                    continue
                ws_candidates = []
                if sheet_no is None:
                    ws_candidates = list(wb.worksheets)
                elif 1 <= sheet_no <= len(wb.worksheets):
                    ws_candidates = [wb.worksheets[sheet_no - 1]]
                for kw in kws:
                    if placed:
                        break
                    for ws in ws_candidates:
                        max_r = int(ws.max_row or 1)
                        if row_no > max_r:
                            continue
                        for c in range(1, min(int(ws.max_column or 1), 128) + 1):
                            if _try_place_at_keyword_cell(ws, row_no, c, kw, sig, dt, max_r):
                                placed = True
                                placed_by = "planned_source_hint"
                                break
                            if placed:
                                break
                        if placed:
                            break
                if placed:
                    break
        if not placed:
            for kw in kws:
                if placed:
                    break
                for ws in wb.worksheets:
                    if placed:
                        break
                    max_r = int(ws.max_row or 1)
                    row_range = (
                        range(max_r, 0, -1) if max_r > 10 else range(1, max_r + 1)
                    )
                    for r in row_range:
                        if placed:
                            break
                        for c in range(1, min(int(ws.max_column or 1), 128) + 1):
                            if _try_place_at_keyword_cell(ws, r, c, kw, sig, dt_use, max_r):
                                placed = True
                                if not placed_by:
                                    placed_by = (
                                        "planned_keywords" if planned_kws else "fallback_keywords"
                                    )
                                break
                            if placed:
                                break
        if isinstance(placement_result, dict):
            one = placement_result.get(str(role_id)) or {}
            one["placed"] = bool(placed)
            if placed:
                one["placed_by"] = placed_by or "fallback_keywords"
            else:
                one["placed_by"] = "not_found"
            placement_result[str(role_id)] = one
    wb.save(out_path)
    return out_path
