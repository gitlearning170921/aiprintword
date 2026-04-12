# -*- coding: utf-8 -*-
"""从 zip（含嵌套 zip）内 docx/xlsx 扫描疑似签字标签短语文本，供维护 sign_role_keywords.json。"""
from __future__ import annotations

import io
import re
import sys
import zipfile
from collections import defaultdict
from typing import BinaryIO, DefaultDict, List, Set

# 命中后才收录，减少噪声
_SIGN_HINT = re.compile(
    r"(?i)\b(author|prepar|review|approv|check|verif|sign|date|signature|endors|validat|accept|release)\b|"
    r"编制|审核|批准|签字|签名|日期|签发|核准|经办|审定|复核|制表|填报|经手|拟稿|起草|会签|质保|质检|授权|确认"
)

_DOCX_TEXT_RE = re.compile(rb"<w:t[^>]*>([^<]*)</w:t>")


def _iter_zip_members(zf: zipfile.ZipFile, prefix: str = "") -> List[tuple[str, zipfile.ZipInfo]]:
    out = []
    for zi in zf.infolist():
        name = prefix + zi.filename
        if zi.is_dir():
            continue
        out.append((name, zi))
    return out


def _extract_docx_text(raw: bytes) -> str:
    parts: List[str] = []
    try:
        with zipfile.ZipFile(io.BytesIO(raw), "r") as dz:
            for n in dz.namelist():
                if not n.endswith(".xml"):
                    continue
                if "word/" not in n.replace("\\", "/"):
                    continue
                try:
                    xml = dz.read(n)
                except Exception:
                    continue
                for m in _DOCX_TEXT_RE.finditer(xml):
                    try:
                        t = m.group(1).decode("utf-8", errors="ignore")
                    except Exception:
                        continue
                    if t.strip():
                        parts.append(t)
    except Exception:
        return ""
    return "\n".join(parts)


def _harvest_docx_strings(raw: bytes, sink: Set[str]) -> None:
    text = _extract_docx_text(raw)
    for line in re.split(r"[\r\n\x0b]+", text):
        s = line.strip()
        if 4 <= len(s) <= 120 and _SIGN_HINT.search(s):
            sink.add(s)


def _harvest_xlsx_strings(raw: bytes, sink: Set[str]) -> None:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(raw), data_only=True)
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is None:
                        continue
                    s = str(v).strip()
                    if 4 <= len(s) <= 120 and _SIGN_HINT.search(s):
                        sink.add(s)
    except Exception:
        pass


def _process_member(name: str, data: bytes, sink: Set[str]) -> None:
    low = name.lower()
    if low.endswith(".docx"):
        _harvest_docx_strings(data, sink)
    elif low.endswith(".xlsx"):
        _harvest_xlsx_strings(data, sink)
    elif low.endswith(".zip"):
        try:
            with zipfile.ZipFile(io.BytesIO(data), "r") as nz:
                scan_zip(nz, sink, prefix=name + "::")
        except Exception:
            pass


def scan_zip(zf: zipfile.ZipFile, sink: Set[str], prefix: str = "") -> None:
    for logical, zi in _iter_zip_members(zf, ""):
        full = prefix + logical
        if zi.filename.endswith("/"):
            continue
        low = zi.filename.lower()
        if not (low.endswith(".docx") or low.endswith(".xlsx") or low.endswith(".zip")):
            continue
        try:
            data = zf.read(zi)
        except Exception:
            continue
        _process_member(full, data, sink)


def classify_role(phrase: str) -> str | None:
    t = phrase.lower()
    # 顺序：更具体的先匹配
    if re.search(
        r"approv|authoriz|signatory|签发|批准|核准", t, re.I
    ) and "review" not in t and "审核" not in t:
        return "approver"
    if re.search(r"review|check|verif|审核|复核|审定|校对", t, re.I):
        return "reviewer"
    if re.search(
        r"author|prepar|draft|编制|作者|拟稿|起草|制表|填报|经办|填表", t, re.I
    ):
        return "author"
    if re.search(r"sign|签字|签名|date|日期", t, re.I):
        return "ambiguous"
    return None


def main() -> int:
    path = (
        sys.argv[1]
        if len(sys.argv) > 1
        else r"c:\Users\yuwell\Downloads\processed_documents (7).zip"
    )
    sink: Set[str] = set()
    with zipfile.ZipFile(path, "r") as zf:
        scan_zip(zf, sink)
    by_role: DefaultDict[str, List[str]] = defaultdict(list)
    amb: List[str] = []
    for s in sorted(sink):
        r = classify_role(s)
        if r == "ambiguous":
            amb.append(s)
        elif r:
            by_role[r].append(s)
        else:
            amb.append(s)
    print("=== by_role ===")
    for k in ("author", "reviewer", "approver"):
        print(k, len(by_role[k]))
        for x in by_role[k]:
            print(" ", repr(x))
    print("=== ambiguous / unclassified ===")
    for x in sorted(amb):
        print(" ", repr(x))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
